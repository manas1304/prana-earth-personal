import os
from typing import Dict, Any, List
import pypdf
import openpyxl

class ESGDocumentParser:
    """
    Parses corporate ESG documentation (PDF Sustainability Reports and Excel BRSR sheets)
    to extract sustainability contexts, emission targets, and resource conservation efforts.
    """
    def __init__(self):
        # Keyword clusters for grouping ESG contexts
        self.keywords = {
            "emissions_carbon": [
                "carbon", "emission", "ghg", "co2", "scope 1", "scope 2", "scope 3",
                "net zero", "decarbonization", "greenhouse gas"
            ],
            "water_management": [
                "water consumption", "water recycled", "water stress", "water harvesting",
                "water usage", "wastewater", "effluent", "groundwater"
            ],
            "energy_transition": [
                "renewable energy", "solar", "wind", "energy efficiency", "biofuel",
                "electricity consumption", "power purchase agreement", "ppa"
            ],
            "biodiversity_land": [
                "biodiversity", "ecosystem", "reforestation", "conservation",
                "deforestation", "habitat", "species", "wildlife"
            ],
            "policies_goals": [
                "brsr", "sustainability policy", "esg goal", "sustainability targets",
                "net-zero target", "climate target", "gri compliance", "disclosure"
            ]
        }

    def parse_sustainability_report_pdf(self, filepath: str) -> Dict[str, Any]:
        """
        Parses a PDF Sustainability Report using pypdf and extracts relevant context snippets.
        """
        results = {
            "file_name": os.path.basename(filepath),
            "file_type": "PDF",
            "total_pages": 0,
            "extracted_text_length": 0,
            "matched_contexts": {cat: [] for cat in self.keywords.keys()},
            "summary": "No data extracted."
        }

        if not os.path.exists(filepath):
            results["summary"] = f"Error: File {filepath} not found."
            return results

        try:
            reader = pypdf.PdfReader(filepath)
            num_pages = len(reader.pages)
            results["total_pages"] = num_pages
            
            full_text_list = []
            for page_num in range(num_pages):
                page = reader.pages[page_num]
                text = page.extract_text()
                if text:
                    full_text_list.append((page_num + 1, text))
            
            total_len = sum(len(t[1]) for t in full_text_list)
            results["extracted_text_length"] = total_len
            
            # Context extraction: search page text for keyword clusters
            for page_num, text in full_text_list:
                lines = text.split("\n")
                for line in lines:
                    line_clean = line.strip().lower()
                    if len(line_clean) < 15: # Skip short lines
                        continue
                    
                    for category, keywords in self.keywords.items():
                        # If a keyword matches, add the line as context
                        for kw in keywords:
                            if kw in line_clean:
                                # Save the context along with page number
                                context_entry = f"[Page {page_num}] {line.strip()}"
                                # Avoid adding exact duplicate lines
                                if context_entry not in results["matched_contexts"][category]:
                                    # Limit to top 10 matches per category for LLM safety
                                    if len(results["matched_contexts"][category]) < 10:
                                        results["matched_contexts"][category].append(context_entry)
                                break

            # Formulate simple summary
            categories_found = [cat for cat, matches in results["matched_contexts"].items() if len(matches) > 0]
            results["summary"] = f"Successfully parsed PDF. Extracted contexts for categories: {', '.join(categories_found)}."
            
        except Exception as e:
            results["summary"] = f"Failed to parse PDF due to exception: {str(e)}"

        return results

    def parse_brsr_disclosure_excel(self, filepath: str) -> Dict[str, Any]:
        """
        Parses an Excel BRSR (Business Responsibility and Sustainability Reporting) sheet
        using openpyxl and extracts target cells matching sustainability topics.
        """
        results = {
            "file_name": os.path.basename(filepath),
            "file_type": "Excel",
            "sheets": [],
            "matched_contexts": {cat: [] for cat in self.keywords.keys()},
            "summary": "No data extracted."
        }

        if not os.path.exists(filepath):
            results["summary"] = f"Error: File {filepath} not found."
            return results

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            results["sheets"] = wb.sheetnames
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Read up to first 200 rows and 20 columns to avoid huge worksheets lagging
                row_idx = 0
                for row in ws.iter_rows(max_row=200, max_col=20, values_only=True):
                    row_idx += 1
                    # Join cell values as strings to search
                    row_values = [str(cell) for cell in row if cell is not None]
                    row_str = " | ".join(row_values).strip()
                    row_str_lower = row_str.lower()
                    
                    if not row_str:
                        continue
                        
                    for category, keywords in self.keywords.items():
                        for kw in keywords:
                            if kw in row_str_lower:
                                context_entry = f"[Sheet: {sheet_name}, Row: {row_idx}] {row_str}"
                                if context_entry not in results["matched_contexts"][category]:
                                    if len(results["matched_contexts"][category]) < 10:
                                        results["matched_contexts"][category].append(context_entry)
                                break
            
            categories_found = [cat for cat, matches in results["matched_contexts"].items() if len(matches) > 0]
            results["summary"] = f"Successfully parsed Excel. Extracted contexts for: {', '.join(categories_found)}."
            
        except Exception as e:
            results["summary"] = f"Failed to parse Excel due to exception: {str(e)}"

        return results
